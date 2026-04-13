# Tests for Excel export.
from pathlib import Path

from openpyxl import load_workbook

from src.export import CSV_COLUMNS
from src.models import BusinessListing
from src.xlsx_export import export_companies_to_xlsx


def test_export_companies_to_xlsx_writes_columns_and_rows(tmp_path: Path) -> None:
    companies = [
        BusinessListing(
            company_name="Acme",
            website="https://acme.test",
            phone="555",
            category="Plumber",
            address="1 St",
            city="Austin, TX",
            source_url="https://yp.test/x",
        )
    ]
    path = tmp_path / "out.xlsx"
    export_companies_to_xlsx(companies, path)
    wb = load_workbook(path)
    ws = wb.active
    assert [c.value for c in ws[1]] == CSV_COLUMNS
    assert ws[2][0].value == "Acme"
